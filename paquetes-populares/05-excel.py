import openpyxl

wb = openpyxl.load_workbook("plantilla.xlsx")
hoja = wb.active

wb.create_sheet("Hoja 3")
sheet3 = wb["Hoja 3"]

sheet3.title = "Hola Mundo"
print(wb.sheetnames)
